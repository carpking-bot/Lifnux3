import json
from datetime import datetime
from pathlib import Path
import openpyxl

ROOT = Path.cwd()
XLSX_PATH = Path(r"C:\Users\dndnjs97\Downloads\DATASET.xlsx")
OUT_PATH = ROOT / "app" / "(apps)" / "finance" / "asset" / "asset_dataset.json"

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws = wb[wb.sheetnames[0]]

START_COL = 3  # C
END_COL = 24   # X

accounts = [
    {"id": "woori_super", "name": "우리SUPER주거래통장", "group": "CASH", "subGroup": "예금/입출금 계좌"},
    {"id": "kb_nara_sarang", "name": "KB나라사랑우대통장", "group": "CASH", "subGroup": "예금/입출금 계좌"},
    {"id": "cash_wallet", "name": "현금", "group": "CASH", "subGroup": "현금"},
    {"id": "kiwoom_kr_active", "name": "키움증권(국내)", "group": "INVESTING", "subGroup": "액티브"},
    {"id": "kiwoom_us_active", "name": "키움증권(해외)", "group": "INVESTING", "subGroup": "액티브"},
    {"id": "namu_isa", "name": "나무증권(ISA)", "group": "INVESTING", "subGroup": "ISA"},
    {"id": "meritz_super365", "name": "메리츠증권(Super365)", "group": "INVESTING", "subGroup": "액티브"},
    {"id": "meritz_pension", "name": "메리츠증권(연금저축펀드)", "group": "INVESTING", "subGroup": "연금저축펀드"},
    {"id": "toss_overseas_active", "name": "토스증권(해외)", "group": "INVESTING", "subGroup": "액티브"},
    {"id": "upbit_crypto", "name": "업비트(코인)", "group": "CASH", "subGroup": "코인"},
    {"id": "naverpay_money", "name": "네이버페이 머니", "group": "CASH", "subGroup": "플랫폼/페이머니"},
    {"id": "kakaopay_money", "name": "카카오페이 머니", "group": "CASH", "subGroup": "플랫폼/페이머니"},
    {"id": "tosspay_money", "name": "토스페이 머니", "group": "CASH", "subGroup": "플랫폼/페이머니"},
    {"id": "nh_housing", "name": "주택청약종합저축(농협)", "group": "SAVING", "subGroup": "주택 청약"},
    {"id": "kb_youth_jump", "name": "KB청년도약계좌", "group": "SAVING", "subGroup": "청약", "memo": "만기 2030.11.03"},
    {"id": "card_woori", "name": "카드값 (우리)", "group": "DEBT", "subGroup": "카드값"},
    {"id": "card_kb", "name": "카드값 (KB)", "group": "DEBT", "subGroup": "카드값"},
]

row_defs = {
    "woori_super": ["우리은행(현금 계좌)"],
    "kb_nara_sarang": ["국민은행(카드값 및 비상금)"],
    "cash_wallet": ["현금"],
    "kiwoom_kr_active": ["키움증권 (국내 액티브 / 예수금)", "키움증권 (국내 액티브 / 잔고)"],
    "kiwoom_us_active": ["키움증권 (해외 액티브 / 예수금)", "키움증권 (해외 액티브 / 잔고)"],
    "namu_isa": ["나무증권 (ISA / 예수금)", "나무증권 (ISA / 잔고)"],
    "meritz_super365": ["메리츠증권 (해외 장기투자 / 예수금)", "메리츠증권 (해외 장기투자 / 잔고)"],
    "meritz_pension": ["메리츠증권 (연금저축펀드 / 예수금)", "메리츠증권 (연금저축펀드  / 잔고)", "메리츠증권 (연금저축펀드 / 잔고)"],
    "toss_overseas_active": ["토스증권 (텐배거 / 예수금)", "토스증권 (텐배거 / 잔고)"],
    "upbit_crypto": ["업비트 (잔고)"],
    "naverpay_money": ["네이버페이 머니"],
    "kakaopay_money": ["카카오페이 머니"],
    "tosspay_money": ["토스페이 머니"],
    "nh_housing": ["주택청약 (농협)"],
    "kb_youth_jump": ["청년도약계좌 (KB)"],
    "card_woori": ["이번달 카드 값(우리)"],
    "card_kb": ["이번달 카드 값(KB)"],
}


def norm_text(v):
    if v is None:
        return ""
    return str(v).strip().replace("\u00a0", " ")


def parse_amount(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(round(v))
    s = norm_text(v)
    if s in {"", "-", "₩ -", "₩-", "￦-"}:
        return 0
    negative = s.startswith("-")
    digits = "".join(ch for ch in s if ch.isdigit())
    if not digits:
        return 0
    n = int(digits)
    return -n if negative else n


def month_from_cell(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m")
    s = norm_text(v).replace(" ", "")
    # 2024.10. / 2024.03 / 2025.1.
    if "." in s:
        parts = [p for p in s.split(".") if p]
        if len(parts) >= 2 and parts[0].isdigit() and parts[1].isdigit():
            return f"{int(parts[0]):04d}-{int(parts[1]):02d}"
    # 2025-01-01 00:00:00
    if "-" in s and len(s) >= 7 and s[:4].isdigit() and s[5:7].isdigit():
        return f"{int(s[:4]):04d}-{int(s[5:7]):02d}"
    raise ValueError(f"Unsupported month cell: {v}")


def created_at_from_cell(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%dT%H:%M:00.000Z")
    s = norm_text(v)
    for fmt in ["%Y.%m.%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"]:
        try:
            dt = datetime.strptime(s, fmt)
            return dt.strftime("%Y-%m-%dT%H:%M:00.000Z")
        except ValueError:
            pass
    # last resort
    return datetime.utcnow().strftime("%Y-%m-%dT%H:%M:00.000Z")

label_to_row = {}
for r in range(1, ws.max_row + 1):
    label = norm_text(ws.cell(r, 2).value)
    if label:
        label_to_row[label] = r

months = [month_from_cell(ws.cell(4, c).value) for c in range(START_COL, END_COL + 1)]
created_ats = [created_at_from_cell(ws.cell(3, c).value) for c in range(START_COL, END_COL + 1)]

expected_totals = []
row_total = label_to_row.get("총 자산")
for c in range(START_COL, END_COL + 1):
    expected_totals.append(parse_amount(ws.cell(row_total, c).value) if row_total else 0)

values_by_account = {}
for acc in accounts:
    series = [0] * len(months)
    for label in row_defs[acc["id"]]:
        r = label_to_row.get(label)
        if not r:
            continue
        for i, c in enumerate(range(START_COL, END_COL + 1)):
            series[i] += parse_amount(ws.cell(r, c).value)
    values_by_account[acc["id"]] = series

snapshots = []
for i, month in enumerate(months):
    lines = [{"accountId": acc["id"], "valueKRW": values_by_account[acc["id"]][i]} for acc in accounts]
    snapshots.append({"month": month, "createdAt": created_ats[i], "lines": lines})

# Requested override:
# add 2024-05 snapshot by cloning 2024-04 values.
if not any(s["month"] == "2024-05" for s in snapshots):
    april_snapshot = next((s for s in snapshots if s["month"] == "2024-04"), None)
    if april_snapshot:
        snapshots.append(
            {
                "month": "2024-05",
                "createdAt": "2024-05-31T23:00:00.000Z",
                "lines": [{"accountId": line["accountId"], "valueKRW": line["valueKRW"]} for line in april_snapshot["lines"]],
            }
        )
        snapshots.sort(key=lambda s: s["month"])
        try:
            april_index = months.index("2024-04")
            expected_totals.insert(april_index + 1, expected_totals[april_index])
        except ValueError:
            expected_totals.insert(1, expected_totals[0] if expected_totals else 0)

mismatches = []
for i, s in enumerate(snapshots):
    actual = sum(line["valueKRW"] for line in s["lines"])
    expected = expected_totals[i]
    if actual != expected:
        mismatches.append({"month": s["month"], "expected": expected, "actual": actual, "diff": actual - expected})

dataset = {
    "accounts": accounts,
    "snapshots": snapshots,
    "validation": {
        "totalAgainstSheet": "OK" if not mismatches else "틀렸습니다",
        "mismatchCount": len(mismatches),
        "mismatches": mismatches,
    },
}

OUT_PATH.write_text(json.dumps(dataset, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"[asset-dataset] accounts={len(accounts)}, snapshots={len(snapshots)}")
print(f"[asset-dataset] first {snapshots[0]['month']} total={sum(l['valueKRW'] for l in snapshots[0]['lines'])}")
print(f"[asset-dataset] last {snapshots[-1]['month']} total={sum(l['valueKRW'] for l in snapshots[-1]['lines'])}")
print(f"[asset-dataset] total check: {dataset['validation']['totalAgainstSheet']} ({dataset['validation']['mismatchCount']})")
